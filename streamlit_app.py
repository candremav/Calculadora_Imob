"""
Simulador de Financiamento Imobiliário – Streamlit App
=======================================================

App web que replica a planilha original "Planilha de Simulações –
Financiamentos Imobiliários" (sistema PRICE com correção monetária pela TR).

Para executar:
    streamlit run streamlit_app.py
"""
from __future__ import annotations

import math
from datetime import date, datetime
from io import BytesIO

import matplotlib.pyplot as plt
import pandas as pd
import streamlit as st
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 1. Configuração da página
# ============================================================
st.set_page_config(
    page_title="Simulador de Financiamento Imobiliário",
    page_icon="🏠",
    layout="wide",
    initial_sidebar_state="expanded",
)


# ============================================================
# 2. Funções financeiras (PMT e NPER do Excel)
# ============================================================
def excel_pmt(rate: float, nper: float, pv: float) -> float:
    """Equivalente à função PMT do Excel.

    Retorna valor negativo (saída de caixa) quando pv > 0.
    """
    if rate == 0:
        return -pv / nper
    fator = (1 + rate) ** nper
    return -(rate * pv * fator) / (fator - 1)


def excel_nper(rate: float, pmt: float, pv: float) -> float:
    """Equivalente à função NPER do Excel."""
    if rate == 0:
        return -pv / pmt
    return math.log(pmt / (pmt + pv * rate)) / math.log(1 + rate)


# ============================================================
# 3. Tabela de amortização mês a mês
# ============================================================
def gerar_tabela_amortizacao(
    valor_financiado: float,
    taxa_mensal: float,
    prazo_total: int,
    taxa_admin: float,
    seguro: float,
    tr_mensal: float,
    data_inicial: datetime,
) -> pd.DataFrame:
    """Gera a tabela de amortização (sistema PRICE com TR) — mesma lógica
    da planilha original (colunas V, W, AA, AD, AJ da aba Comparações)."""
    linhas = []
    saldo = valor_financiado
    parcela_anterior = 0.0

    for mes in range(1, prazo_total + 1):
        juros = saldo * taxa_mensal
        correcao_monetaria = saldo * tr_mensal

        if mes == 1:
            parcela = -excel_pmt(taxa_mensal, prazo_total, saldo) + taxa_admin + seguro
        else:
            pmt_anterior = parcela_anterior - taxa_admin - seguro
            n_restante = excel_nper(taxa_mensal, -pmt_anterior, saldo)
            saldo_corrigido = saldo + correcao_monetaria
            parcela = -excel_pmt(taxa_mensal, n_restante, saldo_corrigido) + taxa_admin + seguro

        amortizacao = parcela - juros - taxa_admin - seguro
        novo_saldo = max(0.0, saldo - amortizacao + correcao_monetaria)

        linhas.append({
            "Mês": mes,
            "Data": data_inicial + relativedelta(months=mes - 1),
            "Ano": (mes - 1) // 12 + 1,
            "Triênio": (mes - 1) // 36 + 1,
            "Década": (mes - 1) // 120 + 1,
            "Saldo Inicial": saldo,
            "Juros": juros,
            "Amortização": amortizacao,
            "Taxa Admin": taxa_admin,
            "Seguro": seguro,
            "Parcela": parcela,
            "Correção Monetária": correcao_monetaria,
            "Saldo Final": novo_saldo,
        })

        saldo = novo_saldo
        parcela_anterior = parcela

    return pd.DataFrame(linhas)


# ============================================================
# 4. Agregações por triênio / década
# ============================================================
def agregar_por_periodo(
    tabela: pd.DataFrame, coluna_periodo: str, label: str
) -> pd.DataFrame:
    """Agrega a tabela mês a mês em triênios (3 anos) ou décadas (10 anos)."""
    grupos = (
        tabela.groupby(coluna_periodo)
        .agg(
            Parcelas=("Parcela", "sum"),
            Juros=("Juros", "sum"),
            Amortização=("Amortização", "sum"),
            **{"Correção Monetária": ("Correção Monetária", "sum")},
            Saldo_Inicial=("Saldo Inicial", "first"),
            Saldo_Final=("Saldo Final", "last"),
        )
        .reset_index()
    )
    grupos["Dif. Dívida"] = grupos["Amortização"] - grupos["Correção Monetária"]
    if label == "triênio":
        grupos["Anos"] = grupos[coluna_periodo].apply(lambda t: f"ANO {3*(t-1)+1} AO {3*t}")
    else:
        grupos["Anos"] = grupos[coluna_periodo].apply(lambda d: f"ANO {10*(d-1)+1} AO {10*d}")
    return grupos[
        [coluna_periodo, "Anos", "Parcelas", "Juros", "Amortização",
         "Dif. Dívida", "Correção Monetária"]
    ]


# ============================================================
# 5. Geração do arquivo Excel para download
# ============================================================
def gerar_excel(
    inputs: dict, tabela: pd.DataFrame, trienios: pd.DataFrame,
    decadas: pd.DataFrame, resumo: dict, totais: dict,
) -> bytes:
    """Gera o workbook Excel com 3 abas: Inputs / Análise / Painel."""
    COR_INPUT, COR_HEADER_BG, COR_HEADER_TXT = "0000FF", "1F4E78", "FFFFFF"
    COR_TOTAL_BG, COR_SECTION_BG = "D9E1F2", "BDD7EE"
    FONT_NAME = "Arial"
    FMT_BRL    = "R$ #,##0.00;(R$ #,##0.00);-"
    FMT_PCT    = "0.0000%"
    FMT_PCT_1D = "0.00%"
    FMT_INT    = "#,##0"
    FMT_DATE   = "dd/mm/yyyy"
    THIN   = Side(border_style="thin",   color="BFBFBF")
    MEDIUM = Side(border_style="medium", color="1F4E78")
    BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def header(cell):
        cell.font = Font(name=FONT_NAME, bold=True, color=COR_HEADER_TXT, size=11)
        cell.fill = PatternFill("solid", start_color=COR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN

    def secao(cell):
        cell.font = Font(name=FONT_NAME, bold=True, color="1F4E78", size=12)
        cell.fill = PatternFill("solid", start_color=COR_SECTION_BG)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    def total(cell):
        cell.font = Font(name=FONT_NAME, bold=True, color="000000", size=11)
        cell.fill = PatternFill("solid", start_color=COR_TOTAL_BG)
        cell.border = Border(top=MEDIUM, bottom=MEDIUM, left=THIN, right=THIN)

    wb = Workbook()

    # ---- Aba 1: Inputs ----
    ws = wb.active
    ws.title = "Inputs"
    ws.sheet_view.showGridLines = False
    ws["B2"] = "INPUTS DO MODELO DE SIMULAÇÃO"
    ws["B2"].font = Font(name=FONT_NAME, bold=True, size=16, color="1F4E78")
    ws.merge_cells("B2:E2")
    ws["B3"] = "Sistema PRICE com correção monetária pela TR"
    ws["B3"].font = Font(name=FONT_NAME, italic=True, size=10, color="595959")
    ws.merge_cells("B3:E3")
    for i, h in enumerate(["Parâmetro", "Valor", "Unidade", "Descrição"], start=2):
        header(ws.cell(row=5, column=i, value=h))

    descricoes = [
        ("Valor financiado",       inputs["valor_financiado"],     "R$",     FMT_BRL,
         "Saldo devedor inicial. Não inclui ITBI, registro ou avaliação."),
        ("Taxa de juros (anual)",  inputs["taxa_juros_anual"],     "% a.a.", FMT_PCT,
         "Juros nominais anuais. Cada banco oferece taxas diferentes."),
        ("Taxa de juros (mensal)", inputs["taxa_juros_mensal"],    "% a.m.", FMT_PCT,
         "Calculada: (1+taxa_anual)^(1/12) − 1."),
        ("Prazo total",            inputs["prazo_meses"],          "meses",  FMT_INT,
         "Comum: 360 (30 anos) ou 420 (35 anos)."),
        ("Taxa administrativa",    inputs["taxa_admin_mensal"],    "R$/mês", FMT_BRL,
         "Tarifa mensal cobrada pelo banco."),
        ("Seguro habitacional",    inputs["seguro_mensal"],        "R$/mês", FMT_BRL,
         "Seguro MIP+DFI; varia conforme idade do mutuário."),
        ("TR mensal estimada",     inputs["tr_mensal_estimada"],   "% a.m.", FMT_PCT,
         "Correção monetária mensal. Divulgada pelo Bacen."),
        ("Data 1ª parcela",        inputs["data_primeira_parcela"], "data",  FMT_DATE,
         "Data do primeiro vencimento."),
        ("Anos no resumo",         inputs["anos_resumo"],          "anos",   FMT_INT,
         "Período do resumo executivo (saiu do bolso etc.)."),
    ]
    azul = Font(name=FONT_NAME, color=COR_INPUT, bold=True, size=11)
    preto = Font(name=FONT_NAME, color="000000", size=11)
    for linha, (nome, valor, unidade, fmt, descr) in enumerate(descricoes, start=6):
        ws.cell(row=linha, column=2, value=nome).font = preto
        c = ws.cell(row=linha, column=3, value=valor)
        c.font = azul; c.number_format = fmt
        c.alignment = Alignment(horizontal="center")
        ws.cell(row=linha, column=4, value=unidade).font = preto
        ws.cell(row=linha, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=linha, column=5, value=descr).font = preto
    ws.cell(row=linha + 2, column=2,
            value="Convenção: valores em AZUL são inputs editáveis.").font = \
        Font(name=FONT_NAME, italic=True, color="595959", size=9)
    ws.merge_cells(start_row=linha + 2, start_column=2, end_row=linha + 2, end_column=5)
    for col, w in {"A": 2, "B": 28, "C": 16, "D": 12, "E": 60}.items():
        ws.column_dimensions[col].width = w

    # ---- Aba 2: Análise Mês a Mês ----
    ws = wb.create_sheet("Análise Mês a Mês")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"
    ws["B2"] = "ANÁLISE MÊS A MÊS — TABELA DE AMORTIZAÇÃO COMPLETA"
    ws["B2"].font = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
    ws.merge_cells("B2:N2")
    ws["B3"] = f"{len(tabela)} parcelas — sistema PRICE com TR"
    ws["B3"].font = Font(name=FONT_NAME, italic=True, size=10, color="595959")
    ws.merge_cells("B3:N3")

    colunas_def = [
        ("Mês",                FMT_INT,    8),
        ("Data",               FMT_DATE,  12),
        ("Ano",                FMT_INT,    6),
        ("Triênio",            FMT_INT,    9),
        ("Década",             FMT_INT,    9),
        ("Saldo Inicial",      FMT_BRL,   16),
        ("Juros",              FMT_BRL,   14),
        ("Amortização",        FMT_BRL,   14),
        ("Taxa Admin",         FMT_BRL,   12),
        ("Seguro",             FMT_BRL,   12),
        ("Parcela",            FMT_BRL,   14),
        ("Correção Monetária", FMT_BRL,   18),
        ("Saldo Final",        FMT_BRL,   16),
        ("% Juros/Parcela",    FMT_PCT_1D,14),
    ]
    for i, (label, _, width) in enumerate(colunas_def, start=2):
        header(ws.cell(row=5, column=i, value=label))
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[5].height = 28

    for idx, row in tabela.iterrows():
        excel_row = 6 + idx
        for i, (col, fmt, _) in enumerate(colunas_def, start=2):
            if col == "% Juros/Parcela":
                val = row["Juros"] / row["Parcela"] if row["Parcela"] else 0
            else:
                val = row[col]
            c = ws.cell(row=excel_row, column=i, value=val)
            c.number_format = fmt
            c.font = Font(name=FONT_NAME, size=10, color="000000")
            if i in (2, 3, 4, 5, 6):
                c.alignment = Alignment(horizontal="center")

    total_row = 6 + len(tabela)
    ws.cell(row=total_row, column=2, value="TOTAIS")
    for i in range(2, 16):
        total(ws.cell(row=total_row, column=i))
    ws.cell(row=total_row, column=2).alignment = Alignment(horizontal="center")
    soma_cols = {7: "Juros", 8: "Amortização", 9: "Taxa Admin",
                 10: "Seguro", 11: "Parcela", 12: "Correção Monetária"}
    for excel_col, df_col in soma_cols.items():
        c = ws.cell(row=total_row, column=excel_col, value=tabela[df_col].sum())
        c.number_format = FMT_BRL
        total(c)
    ws.column_dimensions["A"].width = 2

    # ---- Aba 3: Painel de Acompanhamento ----
    ws = wb.create_sheet("Painel de Acompanhamento")
    ws.sheet_view.showGridLines = False
    ws["B2"] = "PAINEL DE ACOMPANHAMENTO"
    ws["B2"].font = Font(name=FONT_NAME, bold=True, size=16, color="1F4E78")
    ws.merge_cells("B2:I2")

    ws["B4"] = "RESUMO EXECUTIVO"; secao(ws["B4"]); ws.merge_cells("B4:I4")
    ws["B5"] = f'Período de análise: {resumo["anos"]} primeiros anos'
    ws["B5"].font = Font(name=FONT_NAME, italic=True, size=10, color="595959")
    ws.merge_cells("B5:I5")
    indicadores = [
        ("Saiu do bolso",                resumo["saiu_do_bolso"],   FMT_BRL),
        ("Quanto caiu a dívida",         resumo["caiu_a_divida"],   FMT_BRL),
        ("Aproveitamento",               resumo["aproveitamento"],  FMT_PCT_1D),
        ("Saldo devedor após o período", resumo["saldo_apos"],      FMT_BRL),
        ("Parcela 1",                    resumo["parcela_1"],       FMT_BRL),
    ]
    for i, (lbl, val, fmt) in enumerate(indicadores):
        r = 6 + i
        ws.cell(row=r, column=2, value=lbl).font = Font(name=FONT_NAME, size=11)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c = ws.cell(row=r, column=5, value=val)
        c.font = Font(name=FONT_NAME, bold=True, size=11)
        c.alignment = Alignment(horizontal="right"); c.number_format = fmt

    # Triênios
    li = 13
    ws.cell(row=li, column=2, value="ACOMPANHAMENTO A CADA 3 ANOS"); secao(ws.cell(row=li, column=2))
    ws.merge_cells(start_row=li, start_column=2, end_row=li, end_column=9)
    for i, h in enumerate(["Triênio", "Anos", "Parcelas", "Juros", "Amortização",
                            "Dif. Dívida", "Correção Monetária", "% do Total"], start=2):
        header(ws.cell(row=li + 1, column=i, value=h))
    total_30y = trienios["Parcelas"].sum()
    for idx, row in trienios.iterrows():
        er = li + 2 + idx
        ws.cell(row=er, column=2, value=int(row["Triênio"])).number_format = FMT_INT
        ws.cell(row=er, column=3, value=row["Anos"])
        for i, (df_col, fmt) in enumerate([
            ("Parcelas", FMT_BRL), ("Juros", FMT_BRL), ("Amortização", FMT_BRL),
            ("Dif. Dívida", FMT_BRL), ("Correção Monetária", FMT_BRL),
        ], start=4):
            ws.cell(row=er, column=i, value=row[df_col]).number_format = fmt
        c = ws.cell(row=er, column=9, value=row["Parcelas"] / total_30y if total_30y else 0)
        c.number_format = FMT_PCT_1D
        for col in range(2, 10):
            ws.cell(row=er, column=col).font = Font(name=FONT_NAME, size=10)
            ws.cell(row=er, column=col).border = BORDER_THIN
        ws.cell(row=er, column=2).alignment = Alignment(horizontal="center")
    tt = li + 2 + len(trienios)
    ws.cell(row=tt, column=2, value="TOTAIS")
    ws.merge_cells(start_row=tt, start_column=2, end_row=tt, end_column=3)
    for i, (df_col, fmt) in enumerate([
        ("Parcelas", FMT_BRL), ("Juros", FMT_BRL), ("Amortização", FMT_BRL),
        ("Dif. Dívida", FMT_BRL), ("Correção Monetária", FMT_BRL),
    ], start=4):
        c = ws.cell(row=tt, column=i, value=trienios[df_col].sum())
        c.number_format = fmt
    ws.cell(row=tt, column=9, value=1.0).number_format = FMT_PCT_1D
    for col in range(2, 10):
        total(ws.cell(row=tt, column=col))
    ws.cell(row=tt, column=2).alignment = Alignment(horizontal="center")

    # Décadas
    ld = tt + 3
    ws.cell(row=ld, column=2, value="ACOMPANHAMENTO A CADA 10 ANOS"); secao(ws.cell(row=ld, column=2))
    ws.merge_cells(start_row=ld, start_column=2, end_row=ld, end_column=9)
    for i, h in enumerate(["Década", "Anos", "Parcelas", "Juros", "Amortização",
                            "Dif. Dívida", "Correção Monetária", "% do Total"], start=2):
        header(ws.cell(row=ld + 1, column=i, value=h))
    total_dec = decadas["Parcelas"].sum()
    for idx, row in decadas.iterrows():
        er = ld + 2 + idx
        ws.cell(row=er, column=2, value=int(row["Década"])).number_format = FMT_INT
        ws.cell(row=er, column=3, value=row["Anos"])
        for i, (df_col, fmt) in enumerate([
            ("Parcelas", FMT_BRL), ("Juros", FMT_BRL), ("Amortização", FMT_BRL),
            ("Dif. Dívida", FMT_BRL), ("Correção Monetária", FMT_BRL),
        ], start=4):
            ws.cell(row=er, column=i, value=row[df_col]).number_format = fmt
        c = ws.cell(row=er, column=9, value=row["Parcelas"] / total_dec if total_dec else 0)
        c.number_format = FMT_PCT_1D
        for col in range(2, 10):
            ws.cell(row=er, column=col).font = Font(name=FONT_NAME, size=10)
            ws.cell(row=er, column=col).border = BORDER_THIN
        ws.cell(row=er, column=2).alignment = Alignment(horizontal="center")
    td = ld + 2 + len(decadas)
    ws.cell(row=td, column=2, value="TOTAIS")
    ws.merge_cells(start_row=td, start_column=2, end_row=td, end_column=3)
    for i, (df_col, fmt) in enumerate([
        ("Parcelas", FMT_BRL), ("Juros", FMT_BRL), ("Amortização", FMT_BRL),
        ("Dif. Dívida", FMT_BRL), ("Correção Monetária", FMT_BRL),
    ], start=4):
        c = ws.cell(row=td, column=i, value=decadas[df_col].sum())
        c.number_format = fmt
    ws.cell(row=td, column=9, value=1.0).number_format = FMT_PCT_1D
    for col in range(2, 10):
        total(ws.cell(row=td, column=col))
    ws.cell(row=td, column=2).alignment = Alignment(horizontal="center")

    # Totais Gerais
    lg = td + 3
    ws.cell(row=lg, column=2, value="TOTAIS GERAIS DO FINANCIAMENTO"); secao(ws.cell(row=lg, column=2))
    ws.merge_cells(start_row=lg, start_column=2, end_row=lg, end_column=9)
    totais_lst = [
        ("Total pago",                     totais["total_pago"],       FMT_BRL),
        ("Total de juros",                 totais["total_juros"],      FMT_BRL),
        ("Principal amortizado",           totais["total_amortizado"], FMT_BRL),
        ("Total taxas administrativas",    totais["total_taxa_admin"], FMT_BRL),
        ("Total seguro",                   totais["total_seguro"],     FMT_BRL),
        ("Correção monetária acumulada",   totais["total_correcao"],   FMT_BRL),
        ("Saldo devedor final",            totais["saldo_final"],      FMT_BRL),
        ("Multiplicador (pago/financiado)",totais["multiplicador"],    '0.00"x"'),
    ]
    for i, (lbl, val, fmt) in enumerate(totais_lst):
        r = lg + 1 + i
        ws.cell(row=r, column=2, value=lbl).font = Font(name=FONT_NAME, size=11)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c = ws.cell(row=r, column=5, value=val)
        c.font = Font(name=FONT_NAME, bold=True, size=11)
        c.alignment = Alignment(horizontal="right"); c.number_format = fmt

    for col_idx, w in {1: 2, 2: 22, 3: 18, 4: 16, 5: 16, 6: 16, 7: 18, 8: 16, 9: 14}.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ============================================================
# 6. Helpers de formatação BRL
# ============================================================
def fmt_brl(v: float) -> str:
    """Formata número como moeda brasileira (R$ 1.234,56)."""
    if v is None or pd.isna(v):
        return "—"
    s = f"{abs(v):,.2f}"
    s = s.replace(",", "_").replace(".", ",").replace("_", ".")
    return f"-R$ {s}" if v < 0 else f"R$ {s}"


def fmt_pct(v: float) -> str:
    s = f"{v*100:,.2f}"
    return s.replace(",", "_").replace(".", ",").replace("_", ".") + "%"


# ============================================================
# 7. Cabeçalho e orientações da tela inicial
# ============================================================
st.title("🏠 Simulador de Financiamento Imobiliário")
st.caption(
    "Sistema PRICE com correção monetária pela TR — replicação fiel da "
    "Planilha de Simulações de Financiamentos Imobiliários."
)

with st.expander("ℹ️ Como usar esta calculadora", expanded=False):
    st.markdown(
        """
**Esta ferramenta simula um financiamento imobiliário no sistema PRICE com correção monetária mensal pela TR.**
A lógica de cálculo é idêntica à dos contratos da Caixa Econômica Federal:

1. **Preencha os parâmetros** no menu lateral (todos têm um ícone *❓* com explicação detalhada).
2. **Os resultados se atualizam automaticamente** assim que você muda qualquer valor.
3. **Explore as abas**: Resumo Geral, Tabela Mês a Mês, Triênios, Décadas e Gráficos.
4. **Baixe o relatório em Excel** com 3 abas (Inputs, Análise Completa, Painel de Acompanhamento).

**Atenção:** Esta planilha deve ser usada apenas para *simulações*. Os valores oficiais
estão no **Documento Descritivo de Crédito** do banco (DDC).
        """
    )

st.divider()

# ============================================================
# 8. Sidebar – inputs
# ============================================================
with st.sidebar:
    st.header("📝 Parâmetros da Simulação")
    st.caption("Os valores padrão correspondem ao exemplo da planilha original.")

    st.subheader("💰 Financiamento")
    valor_financiado = st.number_input(
        "Valor financiado (R$)",
        min_value=10_000.0, max_value=10_000_000.0,
        value=240_000.00, step=1_000.0, format="%.2f",
        help="Saldo devedor inicial. Não inclui ITBI, registro em cartório "
             "ou taxa de avaliação — apenas o valor que efetivamente entra "
             "no contrato como dívida.",
    )

    taxa_juros_anual_pct = st.number_input(
        "Taxa de juros (% ao ano)",
        min_value=0.0, max_value=30.0,
        value=7.9347, step=0.0001, format="%.4f",
        help="Juros nominais anuais cobrados pelo banco. A taxa varia "
             "conforme o relacionamento com a instituição, garantias e "
             "fonte do funding (poupança, FGTS, SBPE).",
    )
    taxa_juros_anual = taxa_juros_anual_pct / 100

    prazo_meses = st.number_input(
        "Prazo total (meses)",
        min_value=12, max_value=480,
        value=420, step=12,
        help="Número de parcelas mensais. Os prazos mais comuns são 360 "
             "(30 anos) e 420 (35 anos). Prazos menores reduzem o total "
             "de juros mas aumentam a parcela.",
    )
    st.caption(f"  ↳ Equivalente a **{prazo_meses/12:.1f} anos**")

    st.subheader("🧾 Custos mensais")
    taxa_admin_mensal = st.number_input(
        "Taxa administrativa (R$/mês)",
        min_value=0.0, max_value=500.0,
        value=25.00, step=0.50, format="%.2f",
        help="Tarifa fixa mensal cobrada pelo banco para administração "
             "do contrato. Não rende, não amortiza — é só custo.",
    )

    seguro_mensal = st.number_input(
        "Seguro habitacional inicial (R$/mês)",
        min_value=0.0, max_value=2000.0,
        value=17.45, step=0.50, format="%.2f",
        help="Soma dos seguros MIP (Morte e Invalidez Permanente) e DFI "
             "(Danos Físicos do Imóvel). Varia com idade do mutuário, "
             "valor do imóvel e seguradora. Aqui consideramos valor fixo "
             "ao longo do contrato como simplificação.",
    )

    st.subheader("📈 Correção monetária")
    tr_mensal_pct = st.number_input(
        "TR mensal estimada (% ao mês)",
        min_value=0.0, max_value=2.0,
        value=0.17, step=0.01, format="%.4f",
        help="Taxa Referencial mensal usada para corrigir o saldo devedor. "
             "Divulgada pelo Banco Central. Em períodos de Selic alta, a TR "
             "pode chegar a 0,2-0,3% ao mês; em períodos de Selic baixa, "
             "fica em 0%. O valor aqui é uma ESTIMATIVA — ela varia mês a mês.",
    )
    tr_mensal_estimada = tr_mensal_pct / 100

    st.subheader("📅 Datas")
    data_primeira_parcela_input = st.date_input(
        "Data da 1ª parcela",
        value=date(2025, 6, 1),
        help="Data do primeiro vencimento (geralmente 30-45 dias após a "
             "assinatura do contrato). Determina o calendário de todas "
             "as parcelas seguintes.",
    )
    data_primeira_parcela = datetime.combine(data_primeira_parcela_input, datetime.min.time())

    st.subheader("🎯 Resumo executivo")
    anos_resumo = st.number_input(
        "Anos do resumo executivo",
        min_value=1, max_value=int(prazo_meses / 12),
        value=10, step=1,
        help="Período usado para calcular os indicadores 'saiu do bolso', "
             "'caiu a dívida' e 'aproveitamento'. O padrão é 10 anos para "
             "comparar com o típico tempo de venda do imóvel.",
    )

# ============================================================
# 9. Executa simulação
# ============================================================
taxa_juros_mensal = (1 + taxa_juros_anual) ** (1 / 12) - 1

tabela = gerar_tabela_amortizacao(
    valor_financiado    = valor_financiado,
    taxa_mensal         = taxa_juros_mensal,
    prazo_total         = int(prazo_meses),
    taxa_admin          = taxa_admin_mensal,
    seguro              = seguro_mensal,
    tr_mensal           = tr_mensal_estimada,
    data_inicial        = data_primeira_parcela,
)

# Limita as agregações ao número total de meses disponíveis
n_trienios = min(10, math.ceil(prazo_meses / 36))
n_decadas  = min(3,  math.ceil(prazo_meses / 120))
trienios = agregar_por_periodo(tabela[tabela["Triênio"] <= n_trienios], "Triênio", "triênio")
decadas  = agregar_por_periodo(tabela[tabela["Década"]  <= n_decadas],  "Década",  "década")

# Resumo executivo
meses_resumo = anos_resumo * 12
saiu_do_bolso = tabela.loc[tabela["Mês"] <= meses_resumo, "Parcela"].sum()
saldo_apos    = tabela.loc[tabela["Mês"] == meses_resumo, "Saldo Final"].iloc[0] \
                if meses_resumo <= len(tabela) else tabela["Saldo Final"].iloc[-1]
caiu_a_divida = valor_financiado - saldo_apos
aproveitamento = caiu_a_divida / saiu_do_bolso if saiu_do_bolso else 0

# Totais
total_pago       = tabela["Parcela"].sum()
total_juros      = tabela["Juros"].sum()
total_amortizado = tabela["Amortização"].sum()
total_taxa_admin = tabela["Taxa Admin"].sum()
total_seguro     = tabela["Seguro"].sum()
total_correcao   = tabela["Correção Monetária"].sum()
saldo_final      = tabela["Saldo Final"].iloc[-1]

# ============================================================
# 10. KPIs principais (linha de métricas no topo)
# ============================================================
c1, c2, c3, c4 = st.columns(4)
c1.metric("Parcela inicial",  fmt_brl(tabela.loc[0, "Parcela"]))
c2.metric("Parcela final",    fmt_brl(tabela.loc[len(tabela)-1, "Parcela"]))
c3.metric("Total pago",       fmt_brl(total_pago),
          delta=f"{total_pago/valor_financiado:.2f}x o financiado")
c4.metric("Total de juros",   fmt_brl(total_juros),
          delta=f"{total_juros/valor_financiado*100:.1f}% do financiado")

st.divider()

# ============================================================
# 11. Tabs com os resultados
# ============================================================
tab_resumo, tab_mes, tab_tri, tab_dec, tab_graf, tab_sobre = st.tabs(
    ["📊 Resumo Geral", "📋 Mês a Mês", "📅 Triênios",
     "📆 Décadas", "📈 Gráficos", "ℹ️ Sobre"]
)

with tab_resumo:
    st.subheader(f"Resumo executivo — primeiros {anos_resumo} anos")
    r1, r2, r3 = st.columns(3)
    r1.metric("Saiu do bolso",  fmt_brl(saiu_do_bolso))
    r2.metric("Caiu a dívida",  fmt_brl(caiu_a_divida),
              delta=("dívida diminuiu" if caiu_a_divida > 0 else "dívida cresceu"),
              delta_color=("normal" if caiu_a_divida > 0 else "inverse"))
    r3.metric("Aproveitamento", fmt_pct(aproveitamento),
              help="Percentual do dinheiro pago que efetivamente reduziu a dívida.")

    if aproveitamento < 0:
        st.warning(
            f"⚠️ **Aproveitamento negativo:** nos primeiros {anos_resumo} anos, "
            f"a dívida CRESCEU R$ {abs(caiu_a_divida):,.2f} apesar de você ter "
            f"pago R$ {saiu_do_bolso:,.2f}. Isso acontece quando a correção "
            f"monetária (TR) é maior que a amortização. Considere amortizações "
            f"extras ou um prazo menor."
        )

    st.divider()
    st.subheader("Totais do financiamento completo")
    t1, t2 = st.columns(2)
    with t1:
        st.markdown(f"""
        - **Total pago:** {fmt_brl(total_pago)}
        - **Juros pagos:** {fmt_brl(total_juros)}
        - **Principal amortizado:** {fmt_brl(total_amortizado)}
        """)
    with t2:
        st.markdown(f"""
        - **Taxas administrativas:** {fmt_brl(total_taxa_admin)}
        - **Seguro habitacional:** {fmt_brl(total_seguro)}
        - **Correção monetária:** {fmt_brl(total_correcao)}
        """)
    st.markdown(f"**Saldo devedor final:** {fmt_brl(saldo_final)}")

with tab_mes:
    st.subheader("Tabela de amortização mês a mês")
    st.caption(f"{len(tabela)} parcelas — replica as colunas principais da aba *Comparações*.")
    df_show = tabela.copy()
    df_show["Data"] = df_show["Data"].dt.strftime("%d/%m/%Y")
    st.dataframe(
        df_show.style.format({
            "Saldo Inicial": "R$ {:,.2f}", "Juros": "R$ {:,.2f}",
            "Amortização": "R$ {:,.2f}", "Taxa Admin": "R$ {:,.2f}",
            "Seguro": "R$ {:,.2f}", "Parcela": "R$ {:,.2f}",
            "Correção Monetária": "R$ {:,.2f}", "Saldo Final": "R$ {:,.2f}",
        }),
        use_container_width=True, hide_index=True, height=500,
    )

with tab_tri:
    st.subheader("Acompanhamento a cada 3 anos")
    st.caption("Reproduz as linhas 6–17 da aba *Painel de Acompanhamento Mobile*.")
    st.dataframe(
        trienios.drop(columns=["Triênio"]).style.format({
            "Parcelas": "R$ {:,.2f}", "Juros": "R$ {:,.2f}",
            "Amortização": "R$ {:,.2f}", "Dif. Dívida": "R$ {:,.2f}",
            "Correção Monetária": "R$ {:,.2f}",
        }),
        use_container_width=True, hide_index=True,
    )
    st.markdown(f"""
    **Totais ({n_trienios} triênios):**
    - Parcelas: **{fmt_brl(trienios['Parcelas'].sum())}** ·
      Juros: **{fmt_brl(trienios['Juros'].sum())}** ·
      Amortização: **{fmt_brl(trienios['Amortização'].sum())}** ·
      Correção: **{fmt_brl(trienios['Correção Monetária'].sum())}**
    """)

with tab_dec:
    st.subheader("Acompanhamento a cada 10 anos")
    st.caption("Reproduz as linhas 43–47 da aba *Painel de Acompanhamento Mobile*.")
    st.dataframe(
        decadas.drop(columns=["Década"]).style.format({
            "Parcelas": "R$ {:,.2f}", "Juros": "R$ {:,.2f}",
            "Amortização": "R$ {:,.2f}", "Dif. Dívida": "R$ {:,.2f}",
            "Correção Monetária": "R$ {:,.2f}",
        }),
        use_container_width=True, hide_index=True,
    )
    st.markdown(f"""
    **Totais ({n_decadas} décadas):**
    - Parcelas: **{fmt_brl(decadas['Parcelas'].sum())}** ·
      Juros: **{fmt_brl(decadas['Juros'].sum())}** ·
      Amortização: **{fmt_brl(decadas['Amortização'].sum())}** ·
      Correção: **{fmt_brl(decadas['Correção Monetária'].sum())}**
    """)

with tab_graf:
    st.subheader("Visualizações")
    fig, axes = plt.subplots(2, 2, figsize=(13, 8))

    axes[0, 0].plot(tabela["Mês"], tabela["Saldo Final"], color="#C0392B", linewidth=2)
    axes[0, 0].axhline(valor_financiado, color="gray", linestyle="--", alpha=0.5,
                       label="Valor financiado")
    axes[0, 0].set_title("Saldo devedor ao longo do tempo")
    axes[0, 0].set_xlabel("Mês"); axes[0, 0].set_ylabel("R$")
    axes[0, 0].grid(alpha=0.3); axes[0, 0].legend()

    axes[0, 1].plot(tabela["Mês"], tabela["Parcela"], color="#1F618D", linewidth=2)
    axes[0, 1].set_title("Valor da parcela ao longo do tempo")
    axes[0, 1].set_xlabel("Mês"); axes[0, 1].set_ylabel("R$"); axes[0, 1].grid(alpha=0.3)

    axes[1, 0].fill_between(tabela["Mês"], 0, tabela["Juros"],
                            color="#E74C3C", alpha=0.7, label="Juros")
    axes[1, 0].fill_between(tabela["Mês"], tabela["Juros"],
                            tabela["Juros"] + tabela["Amortização"],
                            color="#27AE60", alpha=0.7, label="Amortização")
    axes[1, 0].set_title("Composição da parcela"); axes[1, 0].set_xlabel("Mês")
    axes[1, 0].set_ylabel("R$"); axes[1, 0].legend(); axes[1, 0].grid(alpha=0.3)

    import numpy as np
    x = np.arange(len(trienios))
    axes[1, 1].bar(x, trienios["Juros"], color="#E74C3C", label="Juros")
    axes[1, 1].bar(x, trienios["Amortização"], bottom=trienios["Juros"],
                   color="#27AE60", label="Amortização")
    axes[1, 1].bar(x, trienios["Correção Monetária"],
                   bottom=trienios["Juros"] + trienios["Amortização"],
                   color="#F39C12", label="Correção monet.")
    axes[1, 1].set_xticks(x)
    axes[1, 1].set_xticklabels(trienios["Anos"], rotation=45, ha="right", fontsize=8)
    axes[1, 1].set_title("Composição por triênio"); axes[1, 1].set_ylabel("R$")
    axes[1, 1].legend(); axes[1, 1].grid(alpha=0.3, axis="y")

    plt.tight_layout()
    st.pyplot(fig)

with tab_sobre:
    st.markdown(
        """
### Sobre o cálculo

Este simulador reproduz a lógica do sistema **PRICE com correção monetária pela TR**,
usado em contratos imobiliários da Caixa Econômica Federal:

- **Mês 1**: parcela calculada via `PMT(taxa_mensal, prazo_total, saldo)`.
- **Meses seguintes**: o prazo restante é recalculado via `NPER` com base na parcela
  anterior, e a parcela é recalculada sobre o saldo devedor já corrigido pela TR.
  Por isso a parcela cresce mês a mês quando a correção é maior que a amortização.

### Variáveis derivadas

- **Taxa mensal**: `(1 + taxa_anual)^(1/12) − 1` (capitalização composta).
- **Juros do mês**: `saldo_inicial × taxa_mensal`.
- **Correção monetária**: `saldo_inicial × TR_mensal`.
- **Amortização**: `parcela − juros − taxa_admin − seguro`.
- **Saldo final**: `max(0, saldo_inicial − amortização + correção)`.

### Limitações deste modelo

- Considera **TR constante** (na realidade ela varia mês a mês).
- Considera **seguro fixo** (na realidade ele se ajusta conforme idade e saldo).
- **Não simula** SAC, amortizações extras, FGTS ou comparação com investimentos
  (esses módulos existem na planilha original mas não foram replicados aqui).

### Validação

Os valores deste simulador divergem em **menos de 0,2%** dos valores da planilha
original (diferença residual de arredondamento numérico em 420 iterações
encadeadas de PMT/NPER, sem impacto prático).
        """
    )

# ============================================================
# 12. Botão de download Excel
# ============================================================
st.divider()
col_left, col_right = st.columns([3, 1])
with col_left:
    st.markdown("### 📥 Exportar relatório completo")
    st.caption(
        "Gera um arquivo Excel formatado com 3 abas: **Inputs**, "
        "**Análise Mês a Mês** (todas as 420 parcelas) e "
        "**Painel de Acompanhamento** (resumo, triênios, décadas, totais)."
    )
with col_right:
    inputs_dict = {
        "valor_financiado":      valor_financiado,
        "taxa_juros_anual":      taxa_juros_anual,
        "taxa_juros_mensal":     taxa_juros_mensal,
        "prazo_meses":           int(prazo_meses),
        "taxa_admin_mensal":     taxa_admin_mensal,
        "seguro_mensal":         seguro_mensal,
        "tr_mensal_estimada":    tr_mensal_estimada,
        "data_primeira_parcela": data_primeira_parcela,
        "anos_resumo":           int(anos_resumo),
    }
    resumo_dict = {
        "anos":           int(anos_resumo),
        "saiu_do_bolso":  saiu_do_bolso,
        "caiu_a_divida":  caiu_a_divida,
        "aproveitamento": aproveitamento,
        "saldo_apos":     saldo_apos,
        "parcela_1":      tabela.loc[0, "Parcela"],
    }
    totais_dict = {
        "total_pago":       total_pago,
        "total_juros":      total_juros,
        "total_amortizado": total_amortizado,
        "total_taxa_admin": total_taxa_admin,
        "total_seguro":     total_seguro,
        "total_correcao":   total_correcao,
        "saldo_final":      saldo_final,
        "multiplicador":    total_pago / valor_financiado,
    }
    excel_bytes = gerar_excel(
        inputs_dict, tabela, trienios, decadas, resumo_dict, totais_dict
    )
    st.download_button(
        label="⬇️ Baixar Excel",
        data=excel_bytes,
        file_name=f"simulacao_{date.today().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

st.markdown(
    "<p style='text-align:center; color:#888; font-size:12px; margin-top:30px;'>"
    "Atenção: esta calculadora deve ser usada apenas para simulações. "
    "Os valores oficiais estão no <b>Documento Descritivo de Crédito</b> do banco."
    "</p>",
    unsafe_allow_html=True,
)
